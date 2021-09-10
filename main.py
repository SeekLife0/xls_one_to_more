# coding: utf-8
import sys
import os
import time
import xlrd
import xlwt
import pythoncom
import win32com.client as win32
from xlutils.copy import copy
import openpyxl
reload(sys)
sys.setdefaultencoding('utf-8')

filepaths = []                                  # 初始化列表用来
#获得文件夹下所有文件路径
def all_files_path(rootDir):
    for root, dirs, files in os.walk(rootDir):     # 分别代表根目录、文件夹、文件
        for file in files:                         # 遍历文件
            file_path = os.path.join(root, file)   # 获取文件绝对路径
            filepaths.append(file_path)            # 将文件路径添加进列表
        for dir in dirs:                           # 遍历目录下的子目录
            dir_path = os.path.join(root, dir)     # 获取子目录路径
            all_files_path(dir_path)               # 递归调用

#读取路径下的所有文件名
def read_file_name(file_dir):
    L = []
    for root,dirs,files in os.walk(file_dir):
        for file in files:
            L.append(file)
    return L

# 写入到execel表格
# 先打开原表然后复制一份最后另存为
def wExecel1(path,values,reName):
    oldWb = xlrd.open_workbook(path,formatting_info=True);  # 先打开已存在的表
    newWb = copy(oldWb)  # 复制整个文件,转未xlwt对象这样就可以进行追加写入操作
    #获取所有表
    tables = oldWb.sheets()
    i = 0
    for table in tables:
        if table.nrows == 0:     # 查看表的有效函数是否为0
            continue             # 直接进入下一个循环
        newWbs = newWb.get_sheet(i);  # 取sheet表
        i += 1
        for num in range(0,len(values) - 2,3):
            value = values[num]
            row = values[num+1]
            col = values[num+2]
            print '要写入的坐标行:'+str(row)+'列：'+str(col)
            # 写入的时候添加对应格式,比如居中显示
            style = xlwt.easyxf('font:height 240, color-index black, bold off;align: wrap on, vert centre, horiz center');
            newWbs.write(row, col, value,style)  # 姓名
    # 这里保存的路径需要修改，只需要路径名
    # path路径以点分隔把文件格式后缀删除加上重命名即可
    print path
    print '每个文件重命名的名称'+reName
    eFilePath = path.split('.')[0] + reName + '.xls'
    print '保存路径为：' + eFilePath
    newWb.save(eFilePath)

# 进行xls到xlsx的转换
def xlstoxlsx(filePath):
    if filePath.find('xls'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        filePath1 = filePath.decode('utf-8').encode('gbk')
        wb = None
        try:
            wb = excel.Workbooks.Open(filePath1)
            wb.SaveAs(filePath1 + 'x', FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            print "xls转xlsx的保存路径-->"+filePath + 'x'
        except IOError as e:
            print e
        finally:
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath1 + 'x'

# 进行xlsx到xls的转换
def xlsxtoxls(filePath):
    if filePath.find('xlsx'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        filePath = filePath.decode('utf-8').encode('gbk')
        filePath1 = "{}".format(filePath[:-4])
        wb = None
        try:
            wb = excel.Workbooks.Open(filePath)
            wb.SaveAs(filePath1 + 'xls', FileFormat=56)  # FileFormat = 51 is for .xlsx extension
            print "xlsx转化xls的路径--->"+filePath1.decode('gbk').encode('utf-8') + 'xls'
        except IOError as e:
            print e
        finally:
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath1 + 'xls'

# 使用openpyxl写入到execel表格,这时候转为xlsx才能进行操作
def wExecel3(xlsxPath,values,reName):
    wb = None
    # try:
    wb = openpyxl.load_workbook(xlsxPath)       # 打开表格文件
    sheetnames = wb.get_sheet_names()           # 获取所有表名
    i = 0
    for sheetname in sheetnames:
        sheet = wb.get_sheet_by_name(sheetname)  # 取sheet表
        i += 1
        for num in range(0,len(values) - 2,3):
                value = values[num]
                row = values[num+1]
                col = values[num+2]
                print '写入的值：'+ str(value) + 'row:' + str(row) + 'col:' + str(col)
                sheet.cell(row,col).value = value
        break   # 只写入第一张表
    # 这里保存的路径需要修改，只需要路径名
    # path路径以点分隔把文件格式后缀删除加上重命名即可
    print "开始进行重命名操作的路径-->" + xlsxPath
    print '每个文件重命名的名称'+reName
    eFilePath =  "{}".format(xlsxPath[:-7]) + reName + '.xlsx'   # 命名的魔板文件必须为两个字符
    print '保存路径为：' + eFilePath
    wb.save(eFilePath)

# 因为涉及到一个文件可能存在多个表的问题所以要以表为单位进行内容抓去
def rwExecel_B(Rpath, Wpath, rowValues, reNameCol, eList_entry):
    list_data = []
    xlsPath = Rpath
    xls = None
    reNameC = ''
    # 这里最开始先判断一下填的路径是否存在
    print "判断导入导出路径是否存在"
    if os.path.isfile(Rpath.decode('utf-8').encode('gbk')) and os.path.isfile(Wpath.decode('utf-8').encode('gbk')):
        # 这里需要对路径文件格式进行判断一个if语句
        if Rpath.find('.xlsx') != -1:                    # 把xlsx转为xls，会把原路径的xlsx文件替换
            print '文件格式为：xlsx'
            xlsPath = xlsxtoxls(Rpath)
            # 转换完成之后需要把原来xlsx这个文件删除
            path = Rpath.decode('utf-8').encode('gbk')   # 得到经过转化后的xls文件
            os.remove(path)
            print '文件已删除'
            print "打印输出转化后的路径" + xlsPath.decode('gbk').encode('utf-8')  # 打印输出转化后的路径
        else:
            print '要读取的文件是xls无需转化'
            xlsPath = xlsPath.decode('utf-8').encode('gbk')
            print "无需转化打印输出转化后的路径" + xlsPath.decode('gbk').encode('utf-8')  # 打印输出转化后的路径
        # 因为openpyxl只能操作xlsx所以事先进行转化操作
        # 这里需要对路径文件格式进行判断一个if语句
        if 'xlsx' in Wpath:  # 把xlsx转为xls，会把原路径的xlsx文件替换
            xlsxPath = path.decode('utf-8').encode('gbk')
            print xlsxPath
        else:
            print '文件格式为：xls'
            xlsxPath = xlstoxlsx(Wpath)
        xls = xlrd.open_workbook(xlsPath)
        tables = xls.sheets()           # 获取文件所有表格这样就不需要表的标号和名称了
        sheet_name = xls.sheet_names()
        print sheet_name
        for table in tables:            # 如何判断一个表是否有内容
            if table.nrows == 0:        # 查看表的有效函数是否为0
                continue                # 直接进入下一个循环
            # 通过遍历elist_entry来获得对应坐标,每三个为一个单位获取
            for r in rowValues:
                # 遍历每一列
                for num in range(0,len(eList_entry)-3,4):
                    if (eList_entry[num+1] !='' and eList_entry[num+2]!='' and eList_entry[num+3] != '') and (eList_entry[num+1]!=None and eList_entry[num+2]!=None and eList_entry[num+3]!=None):
                        # 每四个为一组，列名，列数，复制到坐标
                        # 1获取所有的列,第二个是列
                        c = int(eList_entry[num+1])
                        # print reNameCol
                        if c == int(reNameCol):
                            print '重命名的行列' + str(r) + str(c)
                            list_data.append(table.cell_value(r,c))
                            reNameC = table.cell_value(r,c)
                            print '重命名：' + reNameC
                        else:
                            # 获取坐标对应的内容，这里是一整行的内容
                            list_data.append(table.cell_value(r,c))
                        print '每次爬取内容' + str(table.cell_value(r,c))   # 打印内容进行测试
                        # 获取对应要填写的坐标
                        rAim = int(eList_entry[num+2])  # 第三个是坐标行
                        cAim = int(eList_entry[num+3])   # 第四个是坐标列
                        list_data.append(rAim)
                        list_data.append(cAim)
                        # 把所有列遍历完毕，再来进行写入减少io操作
                    else:
                        continue
                wExecel3(xlsxPath, list_data, reNameC)  # 写入的文件路径
                del list_data[:]  # 每次执行完一行清空列表，情况的是列的数据
        # except IOError as e:
        #     print e
    else:
        print "你填入的导出路径不存在请重新填写"

row_values = []
# 模式B的提交按钮处理函数
def deal_Excel_B(grap_row, rename_col, importPath, exportPath):
    print "开始执行-----"
    # 首先对抓去内容进行逗号分割,得到一个列表
    getInputValues = grap_row.split(',')
    # 对此数组进行遍历，查看是否有分隔符-存在
    print '抓取内容的行-->'
    print getInputValues
    for value in getInputValues:
         print '行的范围' + value
         if value.find('-') != -1:
             rages = value.split('-')                     # 继续进行分隔把数据放入列表
             j = int(int(rages[1]) - int(rages[0])) + 1
             print '范围的行数' + str(j)
             for num in range(0, j):
                 num += int(rages[0])                      # 列如：从3开始 7结束
                 print '需要抓取的行：'+ str(num)             # 抓取内容的范围函数
                 row_values.append(num)                    # 添加到行列表中
         else:
             row_values.append(value)
    # 由于xlwt无法成功复制格式，使用openpyxl方式进行xls文件的操作
    rwExecel_B(importPath, exportPath, row_values, rename_col, list_entry)

# 填入对应的参数
grap_row = "1-2"  # 填入列如1,2-4,5 表示抓取1,2,3,4,5行的数据,2-4表示从第2到第4行的所有行
list_entry = ["姓名", "1", "1", "2", "年龄", "2", "1","4", "出生日期", "3", "2", "2"]  # 以四个元素为单位分别是 1列名 2列数 3复制到行 4复制到列 以此类推
rename_col = 1
importPath = "I:/pythonProject/xls_operation_one_to_more/importPath/某机构汇总表.xls"
exportPath = "I:/pythonProject/xls_operation_one_to_more/exportPath/01.xls"

if __name__ == '__main__':
    deal_Excel_B(grap_row, rename_col, importPath, exportPath)